"""Excel report generation for software line mapping.

This module handles generating Excel reports with the mapping results.
"""

import datetime
import logging
from typing import Any, Dict, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from config import TIS_LINK_TEMPLATE

logger = logging.getLogger(__name__)


# Column definitions
MASTER_DATA_HEADERS = [
    "Software Line",
    "ECU - HW Variante",
    "Project class"
]

TIS_STATUS_HEADER = ["Software Line Found in TIS"]

ARTIFACT_HEADERS = [
    "Latest Artifact Found",
    "Project Name",
    "Project RID",
    "Software Line RID",
    "Latest Artifact Name",
    "Latest Artifact RID",
    "Software Type",
    "LCO Version",
    "VeMoX Version",
    "Labcar Type",
    "Life Cycle Status",
    "Upload Path",
    "TIS Link"
]

EXPLANATION_TEXT = [
    "Software Line Mapping Report",
    "{generation_time}",
    "",
    "Purpose:",
    "This report shows the mapping between software lines from the master Excel file "
    "and their corresponding artifacts in TIS.",
    "",
    "Color Coding:",
    "- White columns: Master data from Excel file",
    "- Grey: Software line not found in TIS",
    "- Green: Latest artifact found in TIS",
    "- Red: No artifact found in TIS",
    "",
    "Column Groups:",
    "1. Master Data (white): Original software line information from Excel",
    "2. TIS Status (blue): Indicates if the software line exists in TIS",
    "3. Artifact Data (green): Latest artifact information from TIS",
    "",
    "Notes:",
    "- Software lines are matched flexibly (ignoring spaces, underscores, and special characters)",
    "- TIS links are provided for direct access to the projects",
    ""
]


class ReportGenerator:
    """Generates Excel reports for software line mapping results."""

    def __init__(self):
        self.openpyxl_available = OPENPYXL_AVAILABLE

    def generate_report(
        self, mapping: Dict[str, Any], output_file: str
    ) -> Tuple[bool, Optional[str]]:
        """
        Generate Excel report with mapping results.

        Creates a report with:
        - Explanatory header section
        - Master data (white)
        - Software line existence in TIS (grey if not found)
        - Artifact data (green if found, red if not found)

        Args:
            mapping: The mapping dictionary from MappingHandler
            output_file: Path to save the Excel file

        Returns:
            Tuple of (success, error_message)
        """
        if not self.openpyxl_available:
            return False, "openpyxl is required to generate Excel reports"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Software Line Mapping"

            # Write explanation section
            current_row = self._write_explanation_section(ws)

            # Add separator
            current_row = self._add_separator(ws, current_row)

            # Write headers
            header_row = current_row
            self._write_headers(ws, header_row)

            # Write data rows
            self._write_data_rows(ws, mapping, header_row + 1)

            # Auto-adjust column widths
            self._adjust_column_widths(ws, header_row)

            # Add filter
            self._add_filter(ws, header_row)

            # Save workbook
            wb.save(output_file)
            logger.info(f"Report saved successfully: {output_file}")
            return True, None

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return False, f"Error generating report: {str(e)}"

    def _write_explanation_section(self, ws) -> int:
        """Write the explanation section at the top of the worksheet."""
        explanation_rows = [
            line.format(
                generation_time=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ) if '{generation_time}' in line else line
            for line in EXPLANATION_TEXT
        ]

        max_width = max(len(text) for text in explanation_rows)
        current_row = 1

        for text in explanation_rows:
            cell = ws.cell(row=current_row, column=1, value=text)

            if current_row == 1:  # Title
                cell.font = Font(bold=True, size=14)
                ws.column_dimensions['A'].width = max(max_width * 1.1, 12)
            elif current_row == 2:  # Date
                cell.font = Font(italic=True)
            elif text.endswith(":"):  # Section headers
                cell.font = Font(bold=True)

            current_row += 1

        return current_row

    def _add_separator(self, ws, current_row: int) -> int:
        """Add a separator line and return the new current row."""
        cell = ws.cell(row=current_row, column=1)
        cell.border = Border(bottom=Side(style='double'))
        return current_row + 2

    def _write_headers(self, ws, header_row: int) -> None:
        """Write all header rows."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write master data headers (no background color)
        for col, header in enumerate(MASTER_DATA_HEADERS, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.border = thin_border

        # Write TIS status header (light blue background)
        tis_status_col = len(MASTER_DATA_HEADERS) + 1
        cell = ws.cell(row=header_row, column=tis_status_col, value=TIS_STATUS_HEADER[0])
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.border = thin_border

        # Write artifact headers (with light green background)
        artifact_start_col = len(MASTER_DATA_HEADERS) + 2
        for col, header in enumerate(ARTIFACT_HEADERS, artifact_start_col):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            cell.border = thin_border

    def _write_data_rows(
        self, ws, mapping: Dict[str, Any], start_row: int
    ) -> None:
        """Write all data rows to the worksheet."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        tis_status_col = len(MASTER_DATA_HEADERS) + 1
        artifact_start_col = len(MASTER_DATA_HEADERS) + 2

        current_row = start_row
        for sw_line, data in mapping.items():
            self._write_master_data(ws, current_row, sw_line, data, thin_border)
            self._write_tis_status(ws, current_row, tis_status_col, data, thin_border)
            self._write_artifact_data(
                ws, current_row, artifact_start_col, data, thin_border
            )
            current_row += 1

    def _write_master_data(
        self, ws, row: int, sw_line: str, data: Dict[str, Any], border: Border
    ) -> None:
        """Write master data columns for a single row."""
        master_data = data.get('master_data', {})

        ws.cell(row=row, column=1, value=sw_line)
        ws.cell(row=row, column=2, value=master_data.get("ECU - HW Variante", ""))
        ws.cell(row=row, column=3, value=master_data.get("Project class", ""))

        # Add borders to master data cells
        for col in range(1, len(MASTER_DATA_HEADERS) + 1):
            ws.cell(row=row, column=col).border = border

    def _write_tis_status(
        self, ws, row: int, col: int, data: Dict[str, Any], border: Border
    ) -> None:
        """Write TIS status column for a single row."""
        tis_cell = ws.cell(row=row, column=col, value="Yes" if data['found'] else "No")
        tis_cell.border = border

        if not data['found']:
            tis_cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

    def _write_artifact_data(
        self, ws, row: int, start_col: int, data: Dict[str, Any], border: Border
    ) -> None:
        """Write artifact data columns for a single row."""
        latest_artifact = data.get('latest_artifact', {})
        artifact_found = latest_artifact is not None and bool(latest_artifact)
        fill_color = "E8F5E8" if artifact_found else "FFE6E6"

        col = start_col

        # Basic artifact information
        ws.cell(row=row, column=col, value="Yes" if artifact_found else "No")
        ws.cell(row=row, column=col + 1, value=data['project'])
        ws.cell(row=row, column=col + 2, value=data['project_rid'])
        ws.cell(row=row, column=col + 3, value=data['software_line_rid'])

        if artifact_found:
            ws.cell(row=row, column=col + 4, value=latest_artifact.get('name'))
            artifact_rid = latest_artifact.get('artifact_rid')
            ws.cell(row=row, column=col + 5, value=artifact_rid)
            ws.cell(row=row, column=col + 6, value=latest_artifact.get('software_type'))
            ws.cell(row=row, column=col + 7, value=latest_artifact.get('lco_version'))
            ws.cell(row=row, column=col + 8, value=latest_artifact.get('vemox_version'))
            ws.cell(row=row, column=col + 9, value=latest_artifact.get('labcar_type'))
            ws.cell(row=row, column=col + 10, value=latest_artifact.get('life_cycle_status'))
            ws.cell(row=row, column=col + 11, value=latest_artifact.get('upload_path'))

            # Add TIS link using artifact_rid
            if artifact_rid:
                tis_link = TIS_LINK_TEMPLATE.format(artifact_rid)
                tis_cell = ws.cell(row=row, column=col + 12, value=tis_link)
                tis_cell.hyperlink = tis_link
                tis_cell.style = "Hyperlink"

        # Apply color and borders to artifact data cells
        total_cols = len(MASTER_DATA_HEADERS) + len(ARTIFACT_HEADERS) + 2
        for col_idx in range(start_col, total_cols):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            cell.border = border

    def _adjust_column_widths(self, ws, header_row: int) -> None:
        """Auto-adjust column widths based on content."""
        total_cols = len(MASTER_DATA_HEADERS) + len(ARTIFACT_HEADERS) + 2

        for col_idx in range(1, total_cols):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = max(min(max_length + 2, 70), 12)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_filter(self, ws, header_row: int) -> None:
        """Add auto-filter to the data section."""
        total_cols = len(MASTER_DATA_HEADERS) + len(ARTIFACT_HEADERS) + 1
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(total_cols)}{ws.max_row}"
