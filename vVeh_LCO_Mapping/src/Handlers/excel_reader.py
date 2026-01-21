"""Excel file reading operations.

This module provides functions for reading data from Excel files.
"""

import logging
import traceback
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelReader:
    """Handles reading data from Excel files."""

    def __init__(self):
        self.openpyxl_available = OPENPYXL_AVAILABLE

    def get_excel_data(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> Tuple[Dict[str, Any], Optional[str]]:
        """
        Get comprehensive data from Excel file including software lines and additional columns.

        Args:
            file_path: Path to the Excel file
            sheet_name: Optional specific sheet name to read

        Returns:
            Tuple of (data_dict, error_message)
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.info("Reading Excel file details:")
            logger.info(f"Active sheet: {ws.title}")

            # Define the target columns
            target_columns = {
                "Project line": None,
                "ECU - HW Variante": None,
                "Project class": None
            }

            # Find header row (assuming it's in row 2)
            header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            logger.debug(f"Found headers: {header_row}")

            # Find column indices
            for idx, cell_value in enumerate(header_row):
                if cell_value:
                    cell_str = str(cell_value).strip()
                    logger.debug(f"Checking header: '{cell_str}'")
                    for target_col in target_columns:
                        if cell_str.lower() == target_col.lower():
                            target_columns[target_col] = idx
                            logger.debug(f"Found column '{target_col}' at index {idx}")
                            break

            # Log found column indices
            logger.info("Found column indices:")
            for col, idx in target_columns.items():
                logger.info(f"  {col}: {idx}")

            # Check if we found all required columns
            if target_columns["Project line"] is None:
                wb.close()
                return {}, "Could not find 'Project line' column"

            # Read data
            software_lines = []
            project_data = {}

            logger.debug("Reading data rows...")
            row_count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                row_count += 1
                if not row or not row[target_columns["Project line"]]:
                    continue

                project_line = str(row[target_columns["Project line"]]).strip()
                if project_line and project_line != "Project line":
                    software_lines.append(project_line)

                    # Create data dictionary for this project line
                    hw_variante = (
                        row[target_columns["ECU - HW Variante"]]
                        if target_columns["ECU - HW Variante"] is not None
                        else None
                    )
                    project_class = (
                        row[target_columns["Project class"]]
                        if target_columns["Project class"] is not None
                        else None
                    )

                    project_data[project_line] = {
                        "ECU - HW Variante": str(hw_variante).strip() if hw_variante else "",
                        "Project class": str(project_class).strip() if project_class else ""
                    }

                    # Debug log every 100 rows
                    if row_count % 100 == 0:
                        logger.debug(f"Processed {row_count} rows...")

            wb.close()

            logger.info("Data collection summary:")
            logger.info(f"  Total rows processed: {row_count}")
            logger.info(f"  Software lines found: {len(software_lines)}")
            logger.info(f"  Project data entries: {len(project_data)}")

            # Log sample data at debug level
            if project_data:
                logger.debug("Sample project data (first 3 entries):")
                for i, (key, value) in enumerate(list(project_data.items())[:3]):
                    logger.debug(f"  {key}: {value}")

            result = {
                'software_lines': software_lines,
                'project_data': project_data
            }

            return result, None

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            logger.debug(traceback.format_exc())
            return {}, f"Error reading Excel file: {str(e)}"

    def read_software_lines(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> Tuple[List[str], Optional[str]]:
        """
        Read software lines from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name to read from

        Returns:
            Tuple of (software_lines_list, error_message)
        """
        if not self.openpyxl_available:
            return [], "openpyxl is required to read Excel files"

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.info("Reading Excel file details:")
            logger.info(f"Active sheet: {ws.title}")

            software_lines = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
                if row_idx == 1:  # Header row
                    logger.debug(f"Headers found: {row}")
                    continue

                if row[0]:  # First column
                    value = str(row[0]).strip()
                    if value and value != "Project line":
                        software_lines.append(value)

            wb.close()

            logger.info(f"Found {len(software_lines)} software lines")
            if software_lines:
                logger.debug("First 5 software lines found:")
                for i, line in enumerate(software_lines[:5], 1):
                    logger.debug(f"  {i}. {line}")

            return software_lines, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"

    def get_sheet_names(self, file_path: str) -> Tuple[List[str], Optional[str]]:
        """
        Get all sheet names from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (sheet_names_list, error_message)
        """
        if not self.openpyxl_available:
            return [], "openpyxl is required to read Excel files"

        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"

    def get_column_values_by_header(
        self, file_path: str, header_value: str, sheet_name: Optional[str] = None
    ) -> Tuple[List[str], Optional[str]]:
        """
        Get values from a specific column identified by its header.

        Args:
            file_path: Path to Excel file
            header_value: Header value to search for (e.g., "Project line")
            sheet_name: Sheet name, uses active sheet if None

        Returns:
            Tuple of (values_list, error_message)
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.debug(f"Reading from sheet: {ws.title}")

            # Find the column with the matching header
            header_row = next(ws.iter_rows(min_row=2, max_row=3, values_only=True))
            column_index = None

            logger.debug(f"Looking for header: {header_value}")

            for idx, cell_value in enumerate(header_row):
                if cell_value and str(cell_value).strip().lower() == header_value.lower():
                    column_index = idx
                    logger.debug(f"Found header at column index: {idx}")
                    break

            if column_index is None:
                wb.close()
                return [], f"Header '{header_value}' not found"

            # Read data from the identified column
            data = []
            for row in ws.iter_rows(
                min_row=2,
                min_col=column_index + 1,
                max_col=column_index + 1,
                values_only=True
            ):
                if row[0] is not None:
                    value = str(row[0]).strip()
                    if value and value != "Project line":
                        data.append(value)

            wb.close()

            logger.info(f"Found {len(data)} values")
            if data:
                logger.debug("First 5 values:")
                for i, value in enumerate(data[:5], 1):
                    logger.debug(f"  {i}. {value}")

            return data, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"
