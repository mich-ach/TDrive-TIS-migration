"""Excel report generation for validation results.

This module handles the generation of Excel reports from validation results,
providing multiple sheets for different views of the data.

Functions:
    generate_excel_report: Generate an Excel report with multiple sheets for accountability
"""

import datetime
import logging
from pathlib import Path
from typing import Dict, Optional

from Models import ValidationReport

logger = logging.getLogger(__name__)


def generate_excel_report(
    report: ValidationReport,
    output_dir: Optional[Path] = None,
    component_depth_overrides: Optional[Dict[str, int]] = None
) -> str:
    """
    Generate an Excel report with multiple sheets for accountability.

    Sheets:
    1. Summary - Overall statistics including adaptive depth metrics
    2. All Deviations - Complete list of deviations
    3. By User - Deviations grouped by uploader (accountability)
    4. By Project - Deviations grouped by project
    5. Valid Artifacts - List of correctly placed artifacts
    6. Slow Components - Components that required reduced depth (optional)

    Args:
        report: The ValidationReport containing all validation results
        output_dir: Directory to save the report (defaults to current directory)
        component_depth_overrides: Dict of component IDs to their reduced depth values

    Returns:
        Path to the generated Excel file, or empty string if generation failed
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Skipping Excel report generation.")
        logger.info("Install with: pip install openpyxl")
        return ""

    if output_dir is None:
        output_dir = Path('.')

    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    output_file = output_dir / f"optimized_validation_report_{timestamp}.xlsx"

    wb = Workbook()

    # Define styles
    styles = _create_styles(Font, PatternFill, Border, Side)
    header_font = styles['header_font']
    header_fill = styles['header_fill']
    header_font_white = styles['header_font_white']
    deviation_fill = styles['deviation_fill']
    valid_fill = styles['valid_fill']
    warning_fill = styles['warning_fill']
    info_fill = styles['info_fill']
    thin_border = styles['thin_border']

    # Create sheets
    _create_summary_sheet(wb, report, header_font, component_depth_overrides)
    _create_deviations_sheet(wb, report, header_font_white, header_fill, thin_border,
                             deviation_fill, warning_fill, get_column_letter)
    _create_by_user_sheet(wb, report, header_font_white, header_fill, thin_border, Alignment)
    _create_by_project_sheet(wb, report, header_font_white, header_fill, thin_border, get_column_letter)
    _create_valid_artifacts_sheet(wb, report, header_font_white, thin_border, valid_fill,
                                  PatternFill, get_column_letter)

    if component_depth_overrides:
        _create_slow_components_sheet(wb, component_depth_overrides, header_font_white,
                                      header_fill, thin_border, info_fill)

    # Save workbook
    wb.save(output_file)
    logger.info(f"Excel report saved: {output_file}")

    return str(output_file)


def _create_styles(Font, PatternFill, Border, Side) -> Dict:
    """Create and return all styles used in the report."""
    return {
        'header_font': Font(bold=True, size=11),
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'header_font_white': Font(bold=True, size=11, color="FFFFFF"),
        'deviation_fill': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        'valid_fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'warning_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'info_fill': PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def _create_summary_sheet(wb, report, header_font, component_depth_overrides):
    """Create the Summary sheet."""
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["OPTIMIZED ARTIFACT STRUCTURE VALIDATION REPORT"],
        [f"Generated: {report.timestamp}"],
        [""],
        ["EXECUTION STATISTICS"],
        ["Total Projects", report.total_projects],
        ["Processed Projects", report.processed_projects],
        ["Failed Projects", len(report.failed_projects)],
        ["Total API Calls", report.total_api_calls],
        ["Runtime", f"{report.total_time_seconds:.1f} seconds"],
        [""],
        ["OPTIMIZATION METRICS"],
        ["Cache Hits", report.cache_hits],
        ["Branches Pruned", report.branches_pruned],
        [""],
        ["ADAPTIVE DEPTH METRICS"],
        ["Timeout Retries", report.timeout_retries],
        ["Depth Reductions", report.depth_reductions],
        ["Components with Reduced Depth", len(component_depth_overrides) if component_depth_overrides else 0],
        [""],
        ["ARTIFACT STATISTICS"],
        ["Total Artifacts Found", report.total_artifacts_found],
        ["Valid Artifacts", report.valid_artifacts],
        ["Deviations Found", report.deviations_found],
        [""],
        ["DEVIATIONS BY TYPE"],
    ]

    for dev_type, artifacts in report.deviations_by_type.items():
        if artifacts:
            summary_data.append([dev_type, len(artifacts)])

    summary_data.extend([
        [""],
        ["TOP UPLOADERS WITH DEVIATIONS"],
    ])

    sorted_users = sorted(
        report.deviations_by_user.items(),
        key=lambda x: len(x[1]),
        reverse=True
    )[:10]

    for user, devs in sorted_users:
        summary_data.append([user, len(devs)])

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 4, 11, 15, 20, 25]:
                cell.font = header_font

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20


def _create_deviations_sheet(wb, report, header_font_white, header_fill, thin_border,
                             deviation_fill, warning_fill, get_column_letter):
    """Create the All Deviations sheet."""
    ws_deviations = wb.create_sheet("All Deviations")

    deviation_headers = [
        "Path", "Deviation Type", "Uploader", "Details",
        "Expected Path", "Component ID", "TIS Link"
    ]

    for col_idx, header in enumerate(deviation_headers, 1):
        cell = ws_deviations.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, dev in enumerate(report.deviations, 2):
        ws_deviations.cell(row=row_idx, column=1, value=dev['path']).border = thin_border
        ws_deviations.cell(row=row_idx, column=2, value=dev['deviation_type']).border = thin_border
        ws_deviations.cell(row=row_idx, column=3, value=dev.get('user', 'UNKNOWN')).border = thin_border
        ws_deviations.cell(row=row_idx, column=4, value=dev.get('deviation_details', '')).border = thin_border
        ws_deviations.cell(row=row_idx, column=5, value=dev.get('expected_path_hint', '')).border = thin_border
        ws_deviations.cell(row=row_idx, column=6, value=dev['component_id']).border = thin_border

        tis_link = dev.get('tis_link', '')
        tis_cell = ws_deviations.cell(row=row_idx, column=7, value=tis_link)
        if tis_link:
            tis_cell.hyperlink = tis_link
            tis_cell.style = "Hyperlink"
        tis_cell.border = thin_border

        # Color by deviation type
        fill = warning_fill if dev['deviation_type'] == 'CSP_SWB_UNDER_MODEL' else deviation_fill
        for col in range(1, 8):
            ws_deviations.cell(row=row_idx, column=col).fill = fill

    for col_idx in range(1, len(deviation_headers) + 1):
        ws_deviations.column_dimensions[get_column_letter(col_idx)].width = 30


def _create_by_user_sheet(wb, report, header_font_white, header_fill, thin_border, Alignment):
    """Create the By User (Accountability) sheet."""
    ws_by_user = wb.create_sheet("By User (Accountability)")

    user_headers = ["User", "Total Deviations", "Deviation Types", "Sample Paths"]

    for col_idx, header in enumerate(user_headers, 1):
        cell = ws_by_user.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    all_sorted_users = sorted(
        report.deviations_by_user.items(),
        key=lambda x: len(x[1]),
        reverse=True
    )

    for user, devs in all_sorted_users:
        type_counts = {}
        sample_paths = []
        for d in devs:
            dt = d['deviation_type']
            type_counts[dt] = type_counts.get(dt, 0) + 1
            if len(sample_paths) < 3:
                sample_paths.append(d['path'])

        type_summary = ", ".join([f"{t}: {c}" for t, c in type_counts.items()])
        paths_summary = "\n".join(sample_paths)

        ws_by_user.cell(row=row_idx, column=1, value=user).border = thin_border
        ws_by_user.cell(row=row_idx, column=2, value=len(devs)).border = thin_border
        ws_by_user.cell(row=row_idx, column=3, value=type_summary).border = thin_border

        path_cell = ws_by_user.cell(row=row_idx, column=4, value=paths_summary)
        path_cell.border = thin_border
        path_cell.alignment = Alignment(wrap_text=True)

        row_idx += 1

    ws_by_user.column_dimensions['A'].width = 25
    ws_by_user.column_dimensions['B'].width = 18
    ws_by_user.column_dimensions['C'].width = 45
    ws_by_user.column_dimensions['D'].width = 70


def _create_by_project_sheet(wb, report, header_font_white, header_fill, thin_border, get_column_letter):
    """Create the By Project sheet."""
    ws_by_project = wb.create_sheet("By Project")

    project_headers = ["Project", "Total Deviations", "Uploaders Involved", "Deviation Types"]

    for col_idx, header in enumerate(project_headers, 1):
        cell = ws_by_project.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    for project, devs in sorted(report.deviations_by_project.items(), key=lambda x: -len(x[1])):
        users = set(d.get('user', 'UNKNOWN') for d in devs)
        types = set(d['deviation_type'] for d in devs)

        ws_by_project.cell(row=row_idx, column=1, value=project).border = thin_border
        ws_by_project.cell(row=row_idx, column=2, value=len(devs)).border = thin_border
        ws_by_project.cell(row=row_idx, column=3, value=", ".join(users)).border = thin_border
        ws_by_project.cell(row=row_idx, column=4, value=", ".join(types)).border = thin_border

        row_idx += 1

    for col_idx in range(1, 5):
        ws_by_project.column_dimensions[get_column_letter(col_idx)].width = 35


def _create_valid_artifacts_sheet(wb, report, header_font_white, thin_border, valid_fill,
                                  PatternFill, get_column_letter):
    """Create the Valid Artifacts sheet."""
    ws_valid = wb.create_sheet("Valid Artifacts")

    valid_headers = ["Path", "Uploader", "Component ID", "TIS Link"]

    for col_idx, header in enumerate(valid_headers, 1):
        cell = ws_valid.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = thin_border

    for row_idx, artifact in enumerate(report.valid_paths, 2):
        ws_valid.cell(row=row_idx, column=1, value=artifact['path']).border = thin_border
        ws_valid.cell(row=row_idx, column=2, value=artifact.get('user', 'UNKNOWN')).border = thin_border
        ws_valid.cell(row=row_idx, column=3, value=artifact['component_id']).border = thin_border

        tis_link = artifact.get('tis_link', '')
        tis_cell = ws_valid.cell(row=row_idx, column=4, value=tis_link)
        if tis_link:
            tis_cell.hyperlink = tis_link
            tis_cell.style = "Hyperlink"
        tis_cell.border = thin_border

        for col in range(1, 5):
            ws_valid.cell(row=row_idx, column=col).fill = valid_fill

    for col_idx in range(1, 5):
        ws_valid.column_dimensions[get_column_letter(col_idx)].width = 45


def _create_slow_components_sheet(wb, component_depth_overrides, header_font_white,
                                  header_fill, thin_border, info_fill):
    """Create the Slow Components sheet (optional, for adaptive depth tracking)."""
    ws_slow = wb.create_sheet("Slow Components")

    slow_headers = ["Component ID", "Reduced Depth"]

    for col_idx, header in enumerate(slow_headers, 1):
        cell = ws_slow.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (comp_id, depth) in enumerate(component_depth_overrides.items(), 2):
        ws_slow.cell(row=row_idx, column=1, value=comp_id).border = thin_border
        ws_slow.cell(row=row_idx, column=2, value=depth).border = thin_border

        for col in range(1, 3):
            ws_slow.cell(row=row_idx, column=col).fill = info_fill

    ws_slow.column_dimensions['A'].width = 20
    ws_slow.column_dimensions['B'].width = 15
