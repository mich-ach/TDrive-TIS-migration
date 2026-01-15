"""Handler classes for directory and Excel operations.

This module provides classes for managing directories and Excel file operations.

Classes:
    DirectoryHandler: Handles directory operations for the TIS artifact extraction workflow
    ExcelHandler: Handles Excel file operations for software line mapping
"""

import datetime
import logging
import re
import shutil
import traceback
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import config
from config import TIS_LINK_TEMPLATE

logger = logging.getLogger(__name__)


class DirectoryHandler:
    """Handles directory operations for the TIS artifact extraction workflow."""

    @staticmethod
    def initialize_directories(excel_path: Path) -> Tuple[Path, Path, Path]:
        """
        Initialize output directories and copy Excel file.

        Args:
            excel_path: Path to the input Excel file

        Returns:
            Tuple of (base_output_dir, run_dir, excel_copy_path)

        Raises:
            ValueError: If Excel file not found or directory creation fails
        """
        try:
            # Validate input Excel file
            if not excel_path.exists():
                raise ValueError(f"Excel file not found: {excel_path}")

            # Ensure output directory exists
            config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

            # Create timestamped run directory
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            run_dir = config.OUTPUT_DIR / f"run_{timestamp}"
            run_dir.mkdir(parents=True, exist_ok=True)

            # Copy Excel file to run directory
            excel_copy_path = run_dir / excel_path.name
            shutil.copy2(excel_path, excel_copy_path)

            # Update global config - this is crucial
            config.CURRENT_RUN_DIR = run_dir

            logger.debug(f"Set config.CURRENT_RUN_DIR to: {config.CURRENT_RUN_DIR}")

            return config.OUTPUT_DIR, run_dir, excel_copy_path

        except Exception as e:
            config.CURRENT_RUN_DIR = None  # Reset on failure
            raise ValueError(f"Failed to initialize directories: {str(e)}")

    @staticmethod
    def get_output_file_path(prefix: str, extension: str) -> Path:
        """
        Generate timestamped output file path.

        Args:
            prefix: Prefix for the output file
            extension: File extension (without dot)

        Returns:
            Path object for the output file

        Raises:
            ValueError: If run directory is not initialized
        """
        # Always check the current state from config module
        if not config.CURRENT_RUN_DIR or not config.CURRENT_RUN_DIR.exists():
            raise ValueError(f"Run directory not initialized or doesn't exist! Current value: {config.CURRENT_RUN_DIR}")

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = config.CURRENT_RUN_DIR / f"{prefix}_{timestamp}.{extension}"

        logger.debug(f"Generated output path: {output_path}")

        return output_path

    @staticmethod
    def validate_project_structure() -> bool:
        """
        Validate that all required directories exist or can be created.

        Returns:
            bool: True if structure is valid, False otherwise
        """
        try:
            # Check if source directory exists
            if not config.SCRIPT_DIR.exists():
                logger.error(f"Source directory not found at {config.SCRIPT_DIR}")
                return False

            # Create output directory if it doesn't exist
            config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

            # Verify write permissions
            test_file = config.OUTPUT_DIR / ".test_write"
            try:
                test_file.touch()
                test_file.unlink()
            except Exception as e:
                logger.error(f"No write permission in output directory: {e}")
                return False

            return True

        except Exception as e:
            logger.error(f"Error validating project structure: {e}")
            return False

    @staticmethod
    def cleanup_old_runs(max_runs: int = 999) -> None:
        """
        Clean up old run directories, keeping only the most recent ones.

        Args:
            max_runs: Maximum number of run directories to keep
        """
        try:
            if not config.OUTPUT_DIR.exists():
                return

            # Get all run directories sorted by creation time
            run_dirs = sorted(
                [d for d in config.OUTPUT_DIR.glob("run_*") if d.is_dir()],
                key=lambda x: x.stat().st_mtime,
                reverse=True
            )

            # Remove excess directories
            for old_dir in run_dirs[max_runs:]:
                try:
                    shutil.rmtree(old_dir)
                    logger.info(f"Cleaned up old run directory: {old_dir}")
                except Exception as e:
                    logger.warning(f"Failed to remove old directory {old_dir}: {e}")

        except Exception as e:
            logger.warning(f"Failed to clean up old runs: {e}")

    @staticmethod
    def reset_run_directory() -> None:
        """Reset the current run directory in case of errors."""
        config.CURRENT_RUN_DIR = None

    @staticmethod
    def get_current_run_dir() -> Optional[Path]:
        """Get the current run directory safely."""
        return config.CURRENT_RUN_DIR

    @staticmethod
    def ensure_run_directory_set() -> bool:
        """Ensure run directory is properly set and exists."""
        return (config.CURRENT_RUN_DIR is not None and
                isinstance(config.CURRENT_RUN_DIR, Path) and
                config.CURRENT_RUN_DIR.exists())


class ExcelHandler:
    """Handles Excel file operations for software line mapping."""

    def __init__(self):
        self.openpyxl_available = OPENPYXL_AVAILABLE

    def get_excel_data(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[Dict[str, Any], Optional[str]]:
        """
        Get comprehensive data from Excel file including software lines and additional columns.
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.info("Reading Excel file details:")
            logger.info(f"Active sheet: {ws.title}")

            # Define the target columns
            target_columns = {
                "Project line": None,
                "ECU - HW Variante": None,
                "Project class": None
            }

            # Find header row (assuming it's in row 2)
            header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            logger.debug(f"Found headers: {header_row}")

            # Find column indices
            for idx, cell_value in enumerate(header_row):
                if cell_value:
                    cell_str = str(cell_value).strip()
                    logger.debug(f"Checking header: '{cell_str}'")
                    for target_col in target_columns:
                        if cell_str.lower() == target_col.lower():
                            target_columns[target_col] = idx
                            logger.debug(f"Found column '{target_col}' at index {idx}")
                            break

            # Log found column indices
            logger.info("Found column indices:")
            for col, idx in target_columns.items():
                logger.info(f"  {col}: {idx}")

            # Check if we found all required columns
            if target_columns["Project line"] is None:
                wb.close()
                return {}, "Could not find 'Project line' column"

            # Read data
            software_lines = []
            project_data = {}

            logger.debug("Reading data rows...")
            row_count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                row_count += 1
                if not row or not row[target_columns["Project line"]]:
                    continue

                project_line = str(row[target_columns["Project line"]]).strip()
                if project_line and project_line != "Project line":
                    software_lines.append(project_line)

                    # Create data dictionary for this project line
                    hw_variante = row[target_columns["ECU - HW Variante"]] if target_columns[
                                                                                  "ECU - HW Variante"] is not None else None
                    project_class = row[target_columns["Project class"]] if target_columns[
                                                                                "Project class"] is not None else None

                    project_data[project_line] = {
                        "ECU - HW Variante": str(hw_variante).strip() if hw_variante else "",
                        "Project class": str(project_class).strip() if project_class else ""
                    }

                    # Debug log every 100 rows
                    if row_count % 100 == 0:
                        logger.debug(f"Processed {row_count} rows...")

            wb.close()

            logger.info("Data collection summary:")
            logger.info(f"  Total rows processed: {row_count}")
            logger.info(f"  Software lines found: {len(software_lines)}")
            logger.info(f"  Project data entries: {len(project_data)}")

            # Log sample data at debug level
            if project_data:
                logger.debug("Sample project data (first 3 entries):")
                for i, (key, value) in enumerate(list(project_data.items())[:3]):
                    logger.debug(f"  {key}: {value}")

            result = {
                'software_lines': software_lines,
                'project_data': project_data
            }

            return result, None

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            logger.debug(traceback.format_exc())
            return {}, f"Error reading Excel file: {str(e)}"

    def read_software_lines(self, file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], Optional[str]]:
        """
        Read software lines from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name to read from

        Returns:
            Tuple of (software_lines_list, error_message)
        """
        if not self.openpyxl_available:
            return [], "openpyxl is required to read Excel files"

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.info("Reading Excel file details:")
            logger.info(f"Active sheet: {ws.title}")

            software_lines = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
                if row_idx == 1:  # Header row
                    logger.debug(f"Headers found: {row}")
                    continue

                if row[0]:  # First column
                    value = str(row[0]).strip()
                    if value and value != "Project line":
                        software_lines.append(value)

            wb.close()

            logger.info(f"Found {len(software_lines)} software lines")
            if software_lines:
                logger.debug("First 5 software lines found:")
                for i, line in enumerate(software_lines[:5], 1):
                    logger.debug(f"  {i}. {line}")

            return software_lines, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"

    def get_sheet_names(self, file_path: str) -> Tuple[List[str], Optional[str]]:
        """
        Get all sheet names from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (sheet_names_list, error_message)
        """
        if not self.openpyxl_available:
            return [], "openpyxl is required to read Excel files"

        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"

    def clean_software_line(self, sw_line: str) -> str:
        """
        Clean software line for matching:
        - Remove everything after underscore, space, dash, or opening bracket
        - Remove special characters (dots, etc.)
        - Return uppercase for case-insensitive matching

        Examples:
            "MED17.1.10-6.1" -> "MED17110" (split at dash, remove dots)
            "MG1CS001_test" -> "MG1CS001" (split at underscore)
        """
        if not sw_line:
            return ""

        # First, take everything before underscore, space, dash, or opening bracket
        cleaned = re.split(r'[_\s\-\(\[\{]', sw_line)[0]

        # Remove special characters (dots, etc.)
        cleaned = re.sub(r'[^a-zA-Z0-9]', '', cleaned)

        return cleaned.strip().upper()

    def create_mapping(self, software_lines: List[str], json_data: Dict[str, Any],
                       master_data: Dict[str, Dict[str, str]]) -> Dict[str, Any]:
        """
        Create mapping between software lines and latest artifacts JSON data with flexible matching.
        """
        # Log debug info about master data
        logger.info("Master data info:")
        logger.info(f"  Total entries in master data: {len(master_data)}")
        logger.debug("Sample master data entries:")
        for key in list(master_data.keys())[:3]:
            logger.debug(f"  {key}: {master_data[key]}")

        mapping = {}

        # Create lookup dictionary for json data with cleaned keys
        json_lookup = {}
        for project_name, project_data in json_data.items():
            for sw_line in project_data.get('software_lines', {}):
                cleaned_key = self.clean_software_line(sw_line)
                if cleaned_key:  # Only add if we have a valid cleaned key
                    json_lookup[cleaned_key] = {
                        'original_key': sw_line,
                        'project_name': project_name,
                        'project_data': project_data
                    }

        # Log lookup table for debugging
        if json_data:
            first_project = list(json_data.items())[0][1]
            sw_lines_sample = list(first_project.get('software_lines', {}).keys())[:10]
            logger.debug("Lookup table examples:")
            for original in sw_lines_sample:
                logger.debug(f"  Original: {original} -> Cleaned: {self.clean_software_line(original)}")

        for sw_line in software_lines:
            # Get master data for this software line
            line_master_data = master_data.get(sw_line, {
                "ECU - HW Variante": "",
                "Project class": ""
            })

            mapping[sw_line] = {
                'project': None,
                'project_rid': None,
                'found': False,
                'software_line_rid': None,
                'latest_artifact': None,
                'master_data': line_master_data,
                'matched_with': None
            }

            # Clean the software line for matching
            cleaned_sw_line = self.clean_software_line(sw_line)

            if cleaned_sw_line in json_lookup:
                match_data = json_lookup[cleaned_sw_line]
                original_key = match_data['original_key']
                project_name = match_data['project_name']
                project_data = match_data['project_data']
                sw_line_data = project_data['software_lines'][original_key]

                mapping[sw_line].update({
                    'project': project_name,
                    'project_rid': project_data.get('project_rid'),
                    'found': True,
                    'software_line_rid': sw_line_data.get('software_line_rid'),
                    'latest_artifact': sw_line_data.get('latest_artifact'),
                    'matched_with': original_key
                })

        # Log matching statistics
        matches = sum(1 for m in mapping.values() if m['found'])
        logger.info("Matching Statistics:")
        logger.info(f"  Total software lines: {len(software_lines)}")
        logger.info(f"  Found matches: {matches}")
        logger.info(f"  Missing matches: {len(software_lines) - matches}")

        # Log examples of matches at debug level
        matched_examples = [(sw_line, data['matched_with'])
                            for sw_line, data in mapping.items()
                            if data['found'] and sw_line != data['matched_with']][:5]
        if matched_examples:
            logger.debug("Example matches (original -> matched):")
            for original, matched in matched_examples:
                logger.debug(f"  '{original}' -> '{matched}'")

        return mapping

    def generate_report(self, mapping: Dict[str, Any], output_file: str) -> Tuple[bool, Optional[str]]:
        """
        Generate Excel report with:
        - Explanatory header section
        - Master data (white)
        - Software line existence in TIS (grey if not found)
        - Artifact data (green if found, red if not found)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Software Line Mapping"

            # Add explanatory header section
            explanation_rows = [
                "Software Line Mapping Report",
                f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                "",
                "Purpose:",
                "This report shows the mapping between software lines from the master Excel file and their corresponding artifacts in TIS.",
                "",
                "Color Coding:",
                "- White columns: Master data from Excel file",
                "- Grey: Software line not found in TIS",
                "- Green: Latest artifact found in TIS",
                "- Red: No artifact found in TIS",
                "",
                "Column Groups:",
                "1. Master Data (white): Original software line information from Excel",
                "2. TIS Status (blue): Indicates if the software line exists in TIS",
                "3. Artifact Data (green): Latest artifact information from TIS",
                "",
                "Notes:",
                "- Software lines are matched flexibly (ignoring spaces, underscores, and special characters)",
                "- TIS links are provided for direct access to the projects",
                ""
            ]

            # Write explanation section
            current_row = 1
            max_width = max(len(text) for text in explanation_rows)

            for text in explanation_rows:
                cell = ws.cell(row=current_row, column=1, value=text)
                if current_row == 1:  # Title
                    cell.font = Font(bold=True, size=14)
                elif current_row == 2:  # Date
                    cell.font = Font(italic=True)
                elif text.endswith(":"):  # Section headers
                    cell.font = Font(bold=True)

                if current_row == 1:
                    ws.column_dimensions['A'].width = max(max_width * 1.1, 12)

                current_row += 1

            # Add separator line
            separator_row = current_row
            cell = ws.cell(row=separator_row, column=1)
            cell.border = Border(bottom=Side(style='double'))
            current_row += 2

            # Define headers in correct order
            master_data_headers = [
                "Software Line",
                "ECU - HW Variante",
                "Project class"
            ]

            tis_status_header = ["Software Line Found in TIS"]

            artifact_headers = [
                "Latest Artifact Found",
                "Project Name",
                "Project RID",
                "Software Line RID",
                "Latest Artifact Name",
                "Latest Artifact RID",
                "Software Type",
                "LCO Version",
                "VeMoX Version",
                "Labcar Type",
                "Life Cycle Status",
                "Upload Path",
                "TIS Link"
            ]

            # Write headers
            header_row = current_row

            # Write master data headers (no background color)
            for col, header in enumerate(master_data_headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Write TIS status header (light blue background)
            tis_status_col = len(master_data_headers) + 1
            cell = ws.cell(row=header_row, column=tis_status_col, value=tis_status_header[0])
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write artifact headers (with light green background)
            artifact_start_col = len(master_data_headers) + 2
            for col, header in enumerate(artifact_headers, artifact_start_col):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Write data
            current_row = header_row + 1
            for sw_line, data in mapping.items():
                # Write master data (no background color)
                master_data = data.get('master_data', {})
                ws.cell(row=current_row, column=1, value=sw_line)
                ws.cell(row=current_row, column=2, value=master_data.get("ECU - HW Variante", ""))
                ws.cell(row=current_row, column=3, value=master_data.get("Project class", ""))

                # Add borders to master data cells
                for col in range(1, len(master_data_headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                # Write TIS status with color coding
                tis_cell = ws.cell(row=current_row, column=tis_status_col, value="Yes" if data['found'] else "No")
                tis_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if not data['found']:
                    tis_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                # Write artifact data
                col = artifact_start_col
                latest_artifact = data.get('latest_artifact', {})
                artifact_found = latest_artifact is not None and bool(latest_artifact)
                fill_color = "E8F5E8" if artifact_found else "FFE6E6"

                # Basic artifact information
                ws.cell(row=current_row, column=col, value="Yes" if artifact_found else "No")
                ws.cell(row=current_row, column=col + 1, value=data['project'])
                ws.cell(row=current_row, column=col + 2, value=data['project_rid'])
                ws.cell(row=current_row, column=col + 3, value=data['software_line_rid'])

                if artifact_found:
                    ws.cell(row=current_row, column=col + 4, value=latest_artifact.get('name'))
                    artifact_rid = latest_artifact.get('artifact_rid')
                    ws.cell(row=current_row, column=col + 5, value=artifact_rid)
                    ws.cell(row=current_row, column=col + 6, value=latest_artifact.get('software_type'))
                    ws.cell(row=current_row, column=col + 7, value=latest_artifact.get('lco_version'))
                    ws.cell(row=current_row, column=col + 8, value=latest_artifact.get('vemox_version'))
                    ws.cell(row=current_row, column=col + 9, value=latest_artifact.get('labcar_type'))
                    ws.cell(row=current_row, column=col + 10, value=latest_artifact.get('life_cycle_status'))
                    ws.cell(row=current_row, column=col + 11, value=latest_artifact.get('upload_path'))

                    # Add TIS link using artifact_rid
                    if artifact_rid:
                        tis_link = TIS_LINK_TEMPLATE.format(artifact_rid)
                        tis_cell = ws.cell(row=current_row, column=col + 12, value=tis_link)
                        tis_cell.hyperlink = tis_link
                        tis_cell.style = "Hyperlink"

                # Apply color and borders to artifact data cells
                for col_idx in range(artifact_start_col, len(master_data_headers) + len(artifact_headers) + 2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                current_row += 1

            # Auto-adjust column widths
            for col_idx in range(1, len(master_data_headers) + len(artifact_headers) + 2):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for row in range(header_row, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                adjusted_width = max(min(max_length + 2, 70), 12)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add filter for data section
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(master_data_headers) + len(artifact_headers) + 1)}{ws.max_row}"

            # Save workbook
            wb.save(output_file)
            logger.info(f"Report saved successfully: {output_file}")
            return True, None

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            return False, f"Error generating report: {str(e)}"

    def get_column_values_by_header(self, file_path: str, header_value: str, sheet_name: Optional[str] = None) -> Tuple[
        List[str], Optional[str]]:
        """
        Get values from a specific column identified by its header.

        Args:
            file_path: Path to Excel file
            header_value: Header value to search for (e.g., "Project line")
            sheet_name: Sheet name, uses active sheet if None

        Returns:
            Tuple of (values_list, error_message)
        """
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            logger.debug(f"Reading from sheet: {ws.title}")

            # Find the column with the matching header
            header_row = next(ws.iter_rows(min_row=2, max_row=3, values_only=True))
            column_index = None

            logger.debug(f"Looking for header: {header_value}")

            for idx, cell_value in enumerate(header_row):
                if cell_value and str(cell_value).strip().lower() == header_value.lower():
                    column_index = idx
                    logger.debug(f"Found header at column index: {idx}")
                    break

            if column_index is None:
                wb.close()
                return [], f"Header '{header_value}' not found"

            # Read data from the identified column
            data = []
            for row in ws.iter_rows(min_row=2,
                                    min_col=column_index + 1,
                                    max_col=column_index + 1,
                                    values_only=True):
                if row[0] is not None:
                    value = str(row[0]).strip()
                    if value and value != "Project line":
                        data.append(value)

            wb.close()

            logger.info(f"Found {len(data)} values")
            if data:
                logger.debug("First 5 values:")
                for i, value in enumerate(data[:5], 1):
                    logger.debug(f"  {i}. {value}")

            return data, None

        except Exception as e:
            return [], f"Error reading Excel file: {str(e)}"
