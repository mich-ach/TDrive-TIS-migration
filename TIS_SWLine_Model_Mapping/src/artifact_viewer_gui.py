"""
TIS Artifact Viewer GUI - Simple JSON Data Viewer (wxPython)

A wxPython-based viewer for TIS artifact JSON data.
Loads the complete artifacts file and allows filtering by attributes.

Usage:
    python artifact_viewer_gui.py [json_file]

Requirements:
    pip install wxPython
"""

import json
import webbrowser
from pathlib import Path
from typing import Dict, Any, Optional, List, TYPE_CHECKING
import sys
import datetime

# wxPython import with fallback error message
try:
    import wx  # type: ignore[import-untyped]
    WX_AVAILABLE = True
except ImportError:
    WX_AVAILABLE = False
    print("ERROR: wxPython is not installed.")
    print("Install it with: pip install wxPython")
    print("Note: On Linux, you may need: sudo apt-get install python3-wxgtk4.0")

# openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    import wx  # type: ignore[import-untyped]

# Try to import config for date format, use default if not available
try:
    from config import DATE_DISPLAY_FORMAT, TIS_LINK_TEMPLATE
    OUTPUT_DIR = Path(__file__).parent.parent / "output"
except ImportError:
    DATE_DISPLAY_FORMAT = "%d-%m-%Y %H:%M:%S"
    TIS_LINK_TEMPLATE = "https://rb-ps-tis-dashboard.bosch.com/?gotoCompInstanceId={}"
    OUTPUT_DIR = Path(__file__).parent.parent / "output"


class ArtifactViewerFrame(wx.Frame):  # type: ignore[name-defined]
    """Main frame for TIS Artifact Viewer."""

    def __init__(self, parent: Any, json_file: Optional[Path] = None):
        super().__init__(parent, title="TIS Artifact Viewer", size=(1400, 800))

        self.data: Dict[str, Any] = {}
        self.all_artifacts: List[Dict[str, Any]] = []
        self.filtered_artifacts: List[Dict[str, Any]] = []
        self.current_file: Optional[Path] = None

        self.filter_combos: Dict[str, Any] = {}  # wx.ComboBox instances
        self.search_ctrl: Optional[Any] = None  # wx.TextCtrl instance

        # Sorting state
        self.sort_column: int = -1  # -1 = no sort
        self.sort_ascending: bool = True

        # Track visible columns (non-empty ones)
        self.visible_columns: List[int] = list(range(len(self.columns)))

        # Column definitions: (key, header, min_width, weight, data_key)
        # weight determines how much of extra space the column gets
        # data_key is the artifact dict key to use for this column
        # All columns from JSON data with proper naming
        self.columns = [
            ("project", "Project", 120, 1, "_project"),
            ("sw_line", "Software Line", 120, 1, "_sw_line"),
            ("name", "Name", 180, 3, "name"),
            ("artifact_rid", "Artifact RID", 55, 0, "artifact_rid"),
            ("created_date", "Created Date", 90, 0, "created_date"),
            ("component_type", "Component Type", 75, 1, "component_type"),
            ("simulation_type", "Simulation Type", 70, 0, "simulation_type"),
            ("software_type", "Software Type", 60, 0, "software_type"),
            ("labcar_type", "Labcar Type", 60, 0, "labcar_type"),
            ("test_type", "Test Type", 60, 0, "test_type"),
            ("user", "User", 70, 0, "user"),
            ("lco_version", "LCO Version", 80, 1, "lco_version"),
            ("vemox_version", "Vemox Version", 80, 0, "vemox_version"),
            ("is_genuine_build", "Is Genuine Build", 55, 0, "is_genuine_build"),
            ("life_cycle_status", "Life Cycle Status", 65, 0, "life_cycle_status"),
            ("release_date_time", "Release Date Time", 90, 0, "release_date_time"),
            ("is_deleted", "Is Deleted", 55, 0, "is_deleted"),
            ("deleted_date", "Deleted Date", 80, 0, "deleted_date"),
            ("build_type", "Build Type", 60, 0, "build_type"),
            ("upload_path", "Upload Path", 200, 3, "upload_path"),
        ]

        self._create_ui()

        # Bind resize event to adjust columns
        self.Bind(wx.EVT_SIZE, self._on_resize)

        if json_file:
            wx.CallAfter(self._load_file, json_file)

        self.Centre()

    def _create_ui(self):
        """Create the user interface."""
        # Main panel that fills the frame
        self.panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Toolbar
        toolbar_sizer = wx.BoxSizer(wx.HORIZONTAL)

        open_btn = wx.Button(self.panel, label="Open JSON")
        open_btn.Bind(wx.EVT_BUTTON, self._on_open_file)
        toolbar_sizer.Add(open_btn, 0, wx.ALL, 5)

        latest_btn = wx.Button(self.panel, label="Open Latest")
        latest_btn.Bind(wx.EVT_BUTTON, self._on_open_latest)
        toolbar_sizer.Add(latest_btn, 0, wx.ALL, 5)

        export_btn = wx.Button(self.panel, label="Export to Excel")
        export_btn.Bind(wx.EVT_BUTTON, self._on_export_excel)
        toolbar_sizer.Add(export_btn, 0, wx.ALL, 5)

        self.file_label = wx.StaticText(self.panel, label="No file loaded")
        toolbar_sizer.Add(self.file_label, 0, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 5)

        main_sizer.Add(toolbar_sizer, 0, wx.EXPAND)

        # Filter panel
        filter_box = wx.StaticBox(self.panel, label="Filters (click column headers to sort)")
        filter_sizer = wx.StaticBoxSizer(filter_box, wx.VERTICAL)

        # Row 1: Filter dropdowns
        row1 = wx.BoxSizer(wx.HORIZONTAL)
        self._add_filter(row1, "project", "Project:", 140)
        self._add_filter(row1, "sw_line", "SW Line:", 180)
        self._add_filter(row1, "component_type", "Component:", 120)
        self._add_filter(row1, "simulation_type", "Simulation:", 120)
        self._add_filter(row1, "software_type", "SW Type:", 180)
        self._add_filter(row1, "labcar_type", "Labcar:", 120)
        filter_sizer.Add(row1, 0, wx.EXPAND | wx.ALL, 2)

        # Row 2: More filters
        row2 = wx.BoxSizer(wx.HORIZONTAL)
        self._add_filter(row2, "life_cycle_status", "Status:", 100)
        self._add_filter(row2, "user", "User:", 100)
        self._add_filter(row2, "test_type", "Test Type:", 80)
        self._add_filter(row2, "lco_version", "LCO:", 180)
        self._add_filter(row2, "vemox_version", "VeMoX:", 180)
        self._add_filter(row2, "build_type", "Build Type:", 100)
        filter_sizer.Add(row2, 0, wx.EXPAND | wx.ALL, 2)

        # Row 3: Boolean filters and search
        row3 = wx.BoxSizer(wx.HORIZONTAL)
        self._add_filter(row3, "is_deleted", "Deleted:", 70)
        self._add_filter(row3, "is_genuine_build", "Genuine:", 70)
        row3.Add(wx.StaticText(self.panel, label="Search:"), 0,
                 wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 10)
        self.search_ctrl = wx.TextCtrl(self.panel, size=(150, -1))
        self.search_ctrl.Bind(wx.EVT_TEXT, self._on_filter_changed)
        row3.Add(self.search_ctrl, 0, wx.ALL, 3)
        filter_sizer.Add(row3, 0, wx.EXPAND | wx.ALL, 2)

        # Row 4: Buttons and stats
        row4 = wx.BoxSizer(wx.HORIZONTAL)
        clear_btn = wx.Button(self.panel, label="Clear Filters")
        clear_btn.Bind(wx.EVT_BUTTON, self._on_clear_filters)
        row4.Add(clear_btn, 0, wx.ALL, 3)

        clear_sort_btn = wx.Button(self.panel, label="Clear Sort")
        clear_sort_btn.Bind(wx.EVT_BUTTON, self._on_clear_sort)
        row4.Add(clear_sort_btn, 0, wx.ALL, 3)

        row4.AddStretchSpacer()
        self.sort_label = wx.StaticText(self.panel, label="")
        row4.Add(self.sort_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)
        self.stats_label = wx.StaticText(self.panel, label="")
        row4.Add(self.stats_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 10)

        filter_sizer.Add(row4, 0, wx.EXPAND | wx.ALL, 2)
        main_sizer.Add(filter_sizer, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.TOP, 5)

        # List control - takes remaining space
        self.list_ctrl = wx.ListCtrl(self.panel,
                                     style=wx.LC_REPORT | wx.LC_SINGLE_SEL | wx.BORDER_SUNKEN)

        # Add columns with minimum widths
        for i, (key, header, min_width, weight, data_key) in enumerate(self.columns):
            self.list_ctrl.InsertColumn(i, header, width=min_width)

        # Bind events
        self.list_ctrl.Bind(wx.EVT_LIST_ITEM_ACTIVATED, self._on_item_activated)
        self.list_ctrl.Bind(wx.EVT_LIST_ITEM_RIGHT_CLICK, self._on_context_menu)
        self.list_ctrl.Bind(wx.EVT_LIST_COL_CLICK, self._on_column_click)

        main_sizer.Add(self.list_ctrl, 1, wx.EXPAND | wx.ALL, 5)

        # Set sizer and fit
        self.panel.SetSizer(main_sizer)
        main_sizer.Layout()

        # Status bar
        self.status_bar = self.CreateStatusBar()
        self.status_bar.SetStatusText("Ready")

        # Context menu
        self.context_menu = wx.Menu()
        item_open = self.context_menu.Append(wx.ID_ANY, "Open in TIS")
        item_copy = self.context_menu.Append(wx.ID_ANY, "Copy Artifact RID")
        self.Bind(wx.EVT_MENU, self._on_open_tis, item_open)
        self.Bind(wx.EVT_MENU, self._on_copy_rid, item_copy)

        # Ensure proper layout
        self.Layout()

        # Initial column width adjustment (all columns visible initially)
        wx.CallAfter(self._adjust_column_widths_visible, self.visible_columns)

    def _add_filter(self, sizer: Any, key: str, label: str, width: int,
                    choices: Optional[List[str]] = None):
        """Add a filter combo box."""
        sizer.Add(wx.StaticText(self.panel, label=label), 0,
                  wx.ALIGN_CENTER_VERTICAL | wx.LEFT, 5)

        if choices is None:
            choices = ["All"]

        combo = wx.ComboBox(self.panel, choices=choices, style=wx.CB_READONLY,
                           size=(width, -1))
        combo.SetSelection(0)
        combo.Bind(wx.EVT_COMBOBOX, self._on_filter_changed)
        sizer.Add(combo, 0, wx.ALL, 2)
        self.filter_combos[key] = combo

    def _on_resize(self, event: Any):
        """Handle window resize - adjust column widths proportionally."""
        event.Skip()  # Allow default resize handling

        # Delay column resize to after layout is complete
        wx.CallAfter(self._adjust_column_widths_visible, self.visible_columns)

    def _adjust_column_widths(self):
        """Adjust column widths based on available space."""
        if not hasattr(self, 'list_ctrl') or not self.list_ctrl:
            return

        # Get available width (subtract scrollbar width)
        available_width = self.list_ctrl.GetClientSize().width - 20

        # Calculate total minimum width and total weight
        total_min_width = sum(col[2] for col in self.columns)
        total_weight = sum(col[3] for col in self.columns)

        # Calculate extra space to distribute
        extra_space = max(0, available_width - total_min_width)

        # Set each column width
        for i, (key, header, min_width, weight, data_key) in enumerate(self.columns):
            if total_weight > 0 and extra_space > 0:
                # Distribute extra space proportionally by weight
                extra = int(extra_space * weight / total_weight)
                new_width = min_width + extra
            else:
                new_width = min_width
            self.list_ctrl.SetColumnWidth(i, new_width)

    def _adjust_column_widths_visible(self, visible_cols: List[int]):
        """Adjust column widths based on available space, only for visible columns."""
        if not hasattr(self, 'list_ctrl') or not self.list_ctrl:
            return

        # Get available width (subtract scrollbar width)
        available_width = self.list_ctrl.GetClientSize().width - 20

        # Calculate total minimum width and total weight for visible columns only
        total_min_width = sum(self.columns[i][2] for i in visible_cols)
        total_weight = sum(self.columns[i][3] for i in visible_cols)

        # Calculate extra space to distribute
        extra_space = max(0, available_width - total_min_width)

        # Set each column width
        for i, (key, header, min_width, weight, data_key) in enumerate(self.columns):
            if i in visible_cols:
                if total_weight > 0 and extra_space > 0:
                    # Distribute extra space proportionally by weight
                    extra = int(extra_space * weight / total_weight)
                    new_width = min_width + extra
                else:
                    new_width = min_width
                self.list_ctrl.SetColumnWidth(i, new_width)
            else:
                # Keep hidden columns at 0 width
                self.list_ctrl.SetColumnWidth(i, 0)

    def _on_column_click(self, event: Any):
        """Handle column header click for sorting."""
        col = event.GetColumn()
        if col == self.sort_column:
            # Toggle sort direction
            self.sort_ascending = not self.sort_ascending
        else:
            # New column, sort ascending
            self.sort_column = col
            self.sort_ascending = True

        self._apply_filters()

        # Update sort indicator
        direction = "▲" if self.sort_ascending else "▼"
        col_name = self.columns[col][1]
        self.sort_label.SetLabel(f"Sort: {col_name} {direction}")

    def _on_clear_sort(self, event: Any):
        """Clear sorting."""
        self.sort_column = -1
        self.sort_ascending = True
        self.sort_label.SetLabel("")
        self._apply_filters()

    def _on_open_file(self, event):
        """Open file dialog."""
        initial_dir = str(OUTPUT_DIR) if OUTPUT_DIR.exists() else str(Path.home())

        dlg = wx.FileDialog(
            self, "Select TIS Artifact JSON File",
            defaultDir=initial_dir,
            wildcard="JSON files (*.json)|*.json|All files (*.*)|*.*",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
        )

        if dlg.ShowModal() == wx.ID_OK:
            self._load_file(Path(dlg.GetPath()))
        dlg.Destroy()

    def _on_open_latest(self, event):
        """Open most recent output file."""
        if not OUTPUT_DIR.exists():
            wx.MessageBox(f"Output directory not found:\n{OUTPUT_DIR}",
                         "Warning", wx.OK | wx.ICON_WARNING)
            return

        run_dirs = sorted(OUTPUT_DIR.glob("run_*"), reverse=True)
        if not run_dirs:
            wx.MessageBox("No output runs found.", "Info", wx.OK | wx.ICON_INFORMATION)
            return

        for run_dir in run_dirs:
            json_files = list(run_dir.glob("vveh_lco_artifacts_*.json"))
            if json_files:
                latest_file = max(json_files, key=lambda x: x.stat().st_mtime)
                self._load_file(latest_file)
                return

        wx.MessageBox("No artifact JSON files found.", "Info", wx.OK | wx.ICON_INFORMATION)

    def _on_export_excel(self, event):
        """Export filtered artifacts to Excel."""
        if not OPENPYXL_AVAILABLE:
            wx.MessageBox(
                "openpyxl is not installed.\nInstall it with: pip install openpyxl",
                "Export Error", wx.OK | wx.ICON_ERROR
            )
            return

        if not self.filtered_artifacts:
            wx.MessageBox("No artifacts to export.", "Info", wx.OK | wx.ICON_INFORMATION)
            return

        # Get current filters for filename
        active_filters = []
        for key, combo in self.filter_combos.items():
            value = combo.GetStringSelection()
            if value != "All":
                active_filters.append(f"{key}={value}")

        # Generate default filename
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"artifacts_export_{timestamp}.xlsx"

        # Open save dialog
        dlg = wx.FileDialog(
            self, "Export to Excel",
            defaultDir=str(OUTPUT_DIR) if OUTPUT_DIR.exists() else str(Path.home()),
            defaultFile=default_name,
            wildcard="Excel files (*.xlsx)|*.xlsx",
            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT
        )

        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return

        filepath = Path(dlg.GetPath())
        dlg.Destroy()

        try:
            self._export_to_excel(filepath, active_filters)
            self.status_bar.SetStatusText(f"Exported {len(self.filtered_artifacts)} artifacts to {filepath.name}")
            wx.MessageBox(
                f"Successfully exported {len(self.filtered_artifacts)} artifacts.\n\n"
                f"File: {filepath}",
                "Export Complete", wx.OK | wx.ICON_INFORMATION
            )
        except Exception as e:
            wx.MessageBox(f"Export failed:\n{e}", "Export Error", wx.OK | wx.ICON_ERROR)

    def _export_to_excel(self, filepath: Path, active_filters: List[str]):
        """Export filtered artifacts to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Artifacts"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get non-empty columns only
        non_empty_cols = self._get_non_empty_columns()
        visible_columns = [self.columns[i] for i in non_empty_cols]

        # Write filter info at top
        if active_filters:
            ws['A1'] = "Active Filters:"
            ws['A1'].font = Font(bold=True)
            ws['B1'] = ", ".join(active_filters)
            ws.merge_cells('B1:E1')
            start_row = 3
        else:
            ws['A1'] = "No filters applied (showing all artifacts)"
            ws['A1'].font = Font(italic=True)
            start_row = 3

        # Write headers (only non-empty columns)
        for col_idx, (key, header, min_width, weight, data_key) in enumerate(visible_columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data (only non-empty columns)
        for row_idx, artifact in enumerate(self.filtered_artifacts, start_row + 1):
            for col_idx, (key, header, min_width, weight, data_key) in enumerate(visible_columns, 1):
                value = artifact.get(data_key, '')
                # Format boolean values
                if value is True:
                    value = "Yes"
                elif value is False:
                    value = "No"
                elif value is None:
                    value = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-adjust column widths (only non-empty columns)
        for col_idx, (key, header, min_width, weight, data_key) in enumerate(visible_columns, 1):
            # Calculate max width based on content
            max_length = len(header)
            for row_idx in range(start_row + 1, start_row + 1 + len(self.filtered_artifacts)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            # Cap width at 50 characters
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        summary_ws['A1'] = "Export Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)

        summary_data = [
            ("Export Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source File", self.current_file.name if self.current_file else "N/A"),
            ("Total Artifacts in File", len(self.all_artifacts)),
            ("Filtered Artifacts Exported", len(self.filtered_artifacts)),
            ("", ""),
            ("Active Filters", ""),
        ]

        for row_idx, (label, value) in enumerate(summary_data, 3):
            summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_idx, column=2, value=value)

        # Add filter details
        filter_row = len(summary_data) + 3
        if active_filters:
            for i, f in enumerate(active_filters):
                summary_ws.cell(row=filter_row + i, column=2, value=f)
        else:
            summary_ws.cell(row=filter_row, column=2, value="None (all artifacts shown)")

        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 50

        # Save workbook
        wb.save(filepath)

    def _load_file(self, file_path: Path):
        """Load and parse JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                self.data = json.load(f)

            self.current_file = file_path
            self.file_label.SetLabel(f"File: {file_path.name}")

            self._flatten_artifacts()
            self._update_filter_options()
            self._apply_filters()

            self.status_bar.SetStatusText(f"Loaded: {file_path}")

        except json.JSONDecodeError as e:
            wx.MessageBox(f"Invalid JSON file:\n{e}", "Error", wx.OK | wx.ICON_ERROR)
        except Exception as e:
            wx.MessageBox(f"Failed to load file:\n{e}", "Error", wx.OK | wx.ICON_ERROR)

    def _flatten_artifacts(self):
        """Flatten hierarchical JSON into flat list."""
        self.all_artifacts = []

        for project_name, project_data in self.data.items():
            if not isinstance(project_data, dict):
                continue
            for sw_line_name, sw_line_data in project_data.get('software_lines', {}).items():
                if 'artifacts' in sw_line_data:
                    artifacts = sw_line_data['artifacts']
                elif 'latest_artifact' in sw_line_data:
                    artifact = sw_line_data.get('latest_artifact')
                    artifacts = [artifact] if artifact else []
                else:
                    artifacts = []

                for artifact in artifacts:
                    if artifact and isinstance(artifact, dict):
                        artifact['_project'] = project_name
                        artifact['_sw_line'] = sw_line_name
                        self.all_artifacts.append(artifact)

    def _update_filter_options(self):
        """Update filter dropdown options (initial load - all values)."""
        self._update_dependent_filters(reset_all=True)

    def _update_dependent_filters(self, reset_all: bool = False):
        """Update filter options based on current selections (cascading filters)."""
        # Get current filter selections
        current_selections: Dict[str, str] = {}
        for key, combo in self.filter_combos.items():
            current_selections[key] = combo.GetStringSelection() if not reset_all else "All"

        # Filter artifacts based on "upstream" filters to determine available options
        # Cascade order: Project -> SW Line -> Component -> SW Type -> LCO Version -> Status

        # Start with all artifacts for Project filter
        project_artifacts = self.all_artifacts

        # For SW Line: filter by selected Project
        if current_selections.get('project', 'All') != 'All':
            sw_line_artifacts = [a for a in self.all_artifacts
                                 if str(a.get('_project', '')) == current_selections['project']]
        else:
            sw_line_artifacts = self.all_artifacts

        # For Component/SW Type/LCO/Status: filter by Project AND SW Line
        if current_selections.get('sw_line', 'All') != 'All':
            detail_artifacts = [a for a in sw_line_artifacts
                               if str(a.get('_sw_line', '')) == current_selections['sw_line']]
        else:
            detail_artifacts = sw_line_artifacts

        # Collect unique values for each filter level
        unique_values: Dict[str, set] = {
            'project': set(),
            'sw_line': set(),
            'component_type': set(),
            'simulation_type': set(),
            'software_type': set(),
            'labcar_type': set(),
            'test_type': set(),
            'lco_version': set(),
            'vemox_version': set(),
            'build_type': set(),
            'life_cycle_status': set(),
            'user': set(),
            'is_deleted': set(),
            'is_genuine_build': set(),
        }

        # Project: always show all projects
        for artifact in project_artifacts:
            if artifact.get('_project'):
                unique_values['project'].add(str(artifact['_project']))

        # SW Line: based on selected project
        for artifact in sw_line_artifacts:
            if artifact.get('_sw_line'):
                unique_values['sw_line'].add(str(artifact['_sw_line']))

        # Details: based on project + sw_line selection
        for artifact in detail_artifacts:
            if artifact.get('component_type'):
                unique_values['component_type'].add(str(artifact['component_type']))
            if artifact.get('simulation_type'):
                unique_values['simulation_type'].add(str(artifact['simulation_type']))
            if artifact.get('software_type'):
                unique_values['software_type'].add(str(artifact['software_type']))
            if artifact.get('labcar_type'):
                unique_values['labcar_type'].add(str(artifact['labcar_type']))
            if artifact.get('test_type'):
                unique_values['test_type'].add(str(artifact['test_type']))
            if artifact.get('lco_version'):
                unique_values['lco_version'].add(str(artifact['lco_version']))
            if artifact.get('vemox_version'):
                unique_values['vemox_version'].add(str(artifact['vemox_version']))
            if artifact.get('build_type'):
                unique_values['build_type'].add(str(artifact['build_type']))
            if artifact.get('life_cycle_status'):
                unique_values['life_cycle_status'].add(str(artifact['life_cycle_status']))
            if artifact.get('user'):
                unique_values['user'].add(str(artifact['user']))
            # Handle boolean is_deleted field
            deleted_val = artifact.get('is_deleted')
            if deleted_val is True:
                unique_values['is_deleted'].add("Yes")
            elif deleted_val is False:
                unique_values['is_deleted'].add("No")
            # Handle boolean is_genuine_build field
            genuine_val = artifact.get('is_genuine_build')
            if genuine_val is True:
                unique_values['is_genuine_build'].add("Yes")
            elif genuine_val is False:
                unique_values['is_genuine_build'].add("No")

        # Update each combo box, preserving selection if still valid
        for key, values in unique_values.items():
            if key not in self.filter_combos:
                continue

            combo = self.filter_combos[key]
            current_value = current_selections.get(key, 'All')
            sorted_values = sorted(values)

            combo.Clear()
            combo.Append("All")
            for v in sorted_values:
                combo.Append(v)

            # Restore selection if still valid, otherwise reset to "All"
            # Use index-based selection to preserve case sensitivity
            all_items = ["All"] + sorted_values
            if current_value in all_items:
                idx = all_items.index(current_value)
                combo.SetSelection(idx)
            else:
                combo.SetSelection(0)

    def _on_filter_changed(self, event):
        """Handle filter change."""
        # Update dependent filters first (cascading)
        self._update_dependent_filters()
        # Then apply filters
        self._apply_filters()

    def _on_clear_filters(self, event):
        """Clear all filters."""
        for combo in self.filter_combos.values():
            combo.SetSelection(0)
        if self.search_ctrl:
            self.search_ctrl.SetValue("")
        # Reset all filter options
        self._update_dependent_filters(reset_all=True)
        self._apply_filters()

    def _format_date(self, date_str: str) -> str:
        """Format date string for display."""
        if not date_str:
            return ''
        try:
            if 'T' in date_str:
                clean = date_str.split('.')[0]
                if clean.endswith('Z'):
                    clean = clean[:-1]
                dt = datetime.datetime.fromisoformat(clean)
                return dt.strftime(DATE_DISPLAY_FORMAT)
        except (ValueError, TypeError):
            pass
        return date_str

    def _get_path_without_artifact(self, upload_path: str, artifact_name: str) -> str:
        """Remove artifact name from upload path."""
        if not upload_path:
            return ''
        if artifact_name and upload_path.endswith(artifact_name):
            return upload_path[:-len(artifact_name)].rstrip('/')
        return upload_path

    def _apply_filters(self):
        """Apply all filters, sort, and update the list."""
        filters = {key: combo.GetStringSelection()
                   for key, combo in self.filter_combos.items()}
        search_term = self.search_ctrl.GetValue().lower() if self.search_ctrl else ""

        self.filtered_artifacts = []

        for artifact in self.all_artifacts:
            # Apply filters
            if filters.get('project', 'All') != "All":
                if str(artifact.get('_project', '')) != filters['project']:
                    continue

            if filters.get('sw_line', 'All') != "All":
                if str(artifact.get('_sw_line', '')) != filters['sw_line']:
                    continue

            if filters.get('component_type', 'All') != "All":
                if str(artifact.get('component_type', '') or '') != filters['component_type']:
                    continue

            if filters.get('simulation_type', 'All') != "All":
                if str(artifact.get('simulation_type', '') or '') != filters['simulation_type']:
                    continue

            if filters.get('labcar_type', 'All') != "All":
                if str(artifact.get('labcar_type', '') or '') != filters['labcar_type']:
                    continue

            if filters.get('software_type', 'All') != "All":
                if str(artifact.get('software_type', '') or '') != filters['software_type']:
                    continue

            if filters.get('test_type', 'All') != "All":
                if str(artifact.get('test_type', '') or '') != filters['test_type']:
                    continue

            if filters.get('lco_version', 'All') != "All":
                if str(artifact.get('lco_version', '') or '') != filters['lco_version']:
                    continue

            if filters.get('vemox_version', 'All') != "All":
                if str(artifact.get('vemox_version', '') or '') != filters['vemox_version']:
                    continue

            if filters.get('build_type', 'All') != "All":
                if str(artifact.get('build_type', '') or '') != filters['build_type']:
                    continue

            if filters.get('life_cycle_status', 'All') != "All":
                if str(artifact.get('life_cycle_status', '') or '') != filters['life_cycle_status']:
                    continue

            if filters.get('user', 'All') != "All":
                if str(artifact.get('user', '') or '') != filters['user']:
                    continue

            if filters.get('is_deleted', 'All') != "All":
                deleted_val = artifact.get('is_deleted')
                deleted_str = "Yes" if deleted_val is True else ("No" if deleted_val is False else "")
                if deleted_str != filters['is_deleted']:
                    continue

            if filters.get('is_genuine_build', 'All') != "All":
                genuine_val = artifact.get('is_genuine_build')
                genuine_str = "Yes" if genuine_val is True else ("No" if genuine_val is False else "")
                if genuine_str != filters['is_genuine_build']:
                    continue

            # Search filter
            if search_term:
                searchable = ' '.join([
                    str(artifact.get('name', '')),
                    str(artifact.get('_project', '')),
                    str(artifact.get('_sw_line', '')),
                    str(artifact.get('upload_path', '')),
                    str(artifact.get('user', '') or ''),
                ]).lower()
                if search_term not in searchable:
                    continue

            self.filtered_artifacts.append(artifact)

        # Apply sorting if a column is selected
        if self.sort_column >= 0 and self.sort_column < len(self.columns):
            data_key = self.columns[self.sort_column][4]
            # Date columns need special sorting
            date_columns = {'release_date_time', 'created_date', 'deleted_date'}
            if data_key in date_columns:
                self.filtered_artifacts.sort(
                    key=lambda x: self._parse_date_for_sort(x.get(data_key)),
                    reverse=not self.sort_ascending
                )
            else:
                self.filtered_artifacts.sort(
                    key=lambda x: str(x.get(data_key, '') or '').lower(),
                    reverse=not self.sort_ascending
                )

        # Update display
        self._populate_list()
        self.stats_label.SetLabel(f"Showing {len(self.filtered_artifacts)} of {len(self.all_artifacts)}")

    def _parse_date_for_sort(self, date_str: Optional[str]) -> datetime.datetime:
        """Parse date string for sorting. Returns min datetime for empty/invalid values."""
        if not date_str:
            return datetime.datetime.min
        try:
            # Try DD-MM-YYYY HH:MM:SS format first (DATE_DISPLAY_FORMAT default)
            return datetime.datetime.strptime(date_str, "%d-%m-%Y %H:%M:%S")
        except ValueError:
            try:
                # Try ISO format
                return datetime.datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            except ValueError:
                return datetime.datetime.min

    def _format_cell_value(self, value: Any) -> str:
        """Format a cell value for display."""
        if value is True:
            return "Yes"
        elif value is False:
            return "No"
        elif value is None:
            return ""
        else:
            return str(value)

    def _get_non_empty_columns(self) -> List[int]:
        """Get indices of columns that have at least one non-empty value in filtered artifacts."""
        non_empty_cols = []
        for col_idx, (key, header, min_width, weight, data_key) in enumerate(self.columns):
            has_value = False
            for artifact in self.filtered_artifacts:
                value = artifact.get(data_key)
                # Check if value is non-empty (not None, not empty string, not False for non-boolean)
                if value is not None and value != '':
                    has_value = True
                    break
            if has_value:
                non_empty_cols.append(col_idx)
        return non_empty_cols

    def _populate_list(self):
        """Populate the list control with filtered artifacts."""
        self.list_ctrl.DeleteAllItems()

        # Get columns that have at least one non-empty value
        non_empty_cols = self._get_non_empty_columns()
        self.visible_columns = non_empty_cols  # Store for resize handler

        # Hide empty columns, show non-empty ones
        for col_idx, (key, header, min_width, weight, data_key) in enumerate(self.columns):
            if col_idx in non_empty_cols:
                # Show column with appropriate width
                self.list_ctrl.SetColumnWidth(col_idx, min_width)
            else:
                # Hide column by setting width to 0
                self.list_ctrl.SetColumnWidth(col_idx, 0)

        # Populate rows
        for idx, artifact in enumerate(self.filtered_artifacts):
            # Dynamically populate all columns based on column definitions
            for col_idx, (key, header, min_width, weight, data_key) in enumerate(self.columns):
                value = artifact.get(data_key, '')
                cell_text = self._format_cell_value(value)

                if col_idx == 0:
                    self.list_ctrl.InsertItem(idx, cell_text)
                else:
                    self.list_ctrl.SetItem(idx, col_idx, cell_text)

        # Re-adjust column widths for visible columns
        wx.CallAfter(self._adjust_column_widths_visible, non_empty_cols)

    def _get_selected_artifact(self) -> Optional[Dict]:
        """Get the currently selected artifact."""
        idx = self.list_ctrl.GetFirstSelected()
        if idx != -1 and idx < len(self.filtered_artifacts):
            return self.filtered_artifacts[idx]
        return None

    def _on_item_activated(self, event):
        """Handle double-click on item."""
        self._open_tis_link()

    def _on_context_menu(self, event):
        """Show context menu."""
        self.PopupMenu(self.context_menu)

    def _on_open_tis(self, event):
        """Open TIS link for selected artifact."""
        self._open_tis_link()

    def _open_tis_link(self):
        """Open TIS link."""
        artifact = self._get_selected_artifact()
        if artifact and artifact.get('artifact_rid'):
            url = TIS_LINK_TEMPLATE.format(artifact['artifact_rid'])
            webbrowser.open(url)

    def _on_copy_rid(self, event):
        """Copy RID to clipboard."""
        artifact = self._get_selected_artifact()
        if artifact and artifact.get('artifact_rid'):
            rid = str(artifact['artifact_rid'])
            if wx.TheClipboard.Open():
                wx.TheClipboard.SetData(wx.TextDataObject(rid))
                wx.TheClipboard.Close()
                self.status_bar.SetStatusText(f"Copied: {rid}")


def main():
    """Main entry point."""
    if not WX_AVAILABLE:
        sys.exit(1)

    app = wx.App(False)  # type: ignore[name-defined]

    json_file = None
    if len(sys.argv) > 1:
        json_file = Path(sys.argv[1])
        if not json_file.exists():
            print(f"File not found: {json_file}")
            json_file = None

    frame = ArtifactViewerFrame(None, json_file)
    frame.Show()
    app.MainLoop()


if __name__ == "__main__":
    main()
